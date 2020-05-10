from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def check(data, pnum, url, aboveRank=False):
    fill = PatternFill("solid", fgColor="00FF00")
    gfill = PatternFill("solid", fgColor="FF0000")

    workbook = load_workbook(url)
    result = Workbook()
    rSheet = result.active
    rNow = 0

    sheets = workbook.get_sheet_names()
    bookSheet = workbook.get_sheet_by_name(sheets[0])

    rows = bookSheet.rows
    first = True

    for row in rows:

        if row[0].value is None:
            break
        if first:
            rNow += 1
            for col in range(1, 7):
                if col == 1:
                    val = '名字'
                elif col == 2:
                    val = '用户名'
                elif col == 3:
                    val = '出题数'
                elif col == 4:
                    val = '罚时'
                elif col == 5:
                    val = '补题数'
                else:
                    val = '题目提交数'
                rSheet.cell(rNow, col).value = val
            for col in range(pnum):
                rSheet.cell(rNow, 7 + col).value = chr(65 + col)
            if aboveRank:
                rSheet.cell(rNow, 7 + pnum).value = "Rank"
            first = False
            continue

        rNow += 1
        who = str(row[1].value)
        if who not in data or not data[who]['attempted']:
            rSheet.cell(rNow, 1).fill = gfill
            rSheet.cell(rNow, 1).value = row[0].value
            rSheet.cell(rNow, 2).fill = gfill
            rSheet.cell(rNow, 2).value = '缺席原因'
            if who not in data:
                continue

        for col in range(1, 6):
            if col == 1:
                val = row[0].value
            elif col == 2:
                val = row[1].value
            elif col == 3:
                val = len(data[who]['solved'])
            elif col == 4:
                val = data[who]['totTime'] // 60
            else:
                val = len(data[who]['solved']) + len(data[who]['aSolved'])
                if val == pnum:
                    rSheet.cell(rNow, col).fill = fill
            rSheet.cell(rNow, col).value = val
        for col in range(pnum):
            if col in data[who]['solved'] or col in data[who]['aSolved']:
                solve = 1
            else:
                solve = 0
            if col in data[who]:
                rSheet.cell(rNow, 7 + col).value = str(data[who][col] + solve) + '/' + str(solve)
            else:
                rSheet.cell(rNow, 7 + col).value = '0/0'
        if aboveRank:
            rSheet.cell(rNow, 7 + pnum).value = data[who]['rank']

    return result
